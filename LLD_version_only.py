import os
import re
import threading
import queue
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from rapidfuzz import fuzz


# =========================
# Defaults / Constants
# =========================
FALLBACK_CMD_MIN = 0.85
LOW_SCORE_THRESHOLD = 0.15

END_MARKER_RE = re.compile(r"\[END\]\s*$", re.IGNORECASE)
CMD_LINE_RE = re.compile(r"^\s*#\s*(.+?)\s*$")

HDR_NAME = {"name"}
HDR_CMD = {"command", "cmd"}
HDR_IN = {"input data", "input", "in", "inputdata"}
HDR_OUT = {"output data", "output", "out", "outputdata"}
HDR_SOURCE = {"source"}

GENERIC_CMD_TOKENS = {
    "oc", "kubectl", "get", "pod", "pods", "po", "describe", "apply", "delete",
    "docker", "images", "grep", "jq", "awk", "head", "xargs", "cut", "uniq", "sort", "sed", "tr",
    "-n", "--namespace", "-o", "-oname", "-ojson", "-oyaml", "-o<out>",
}

STOP_WORDS = {"bash", "error", "help", "see", "use"}


# =========================
# Helpers
# =========================
def now_ts() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def compact_ws(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip())


def strip_quotes_unescape(s: str) -> str:
    s = s.replace('\\"', '"').replace("\\'", "'")
    s = s.replace("“", '"').replace("”", '"').replace("‘", "'").replace("’", "'")
    return s


def normalize_command(cmd: str) -> str:
    if cmd is None:
        return ""
    s = strip_quotes_unescape(str(cmd)).strip()
    s = re.sub(r"\s*\|\s*", "|", s)
    s = compact_ws(s)

    # limited idiomatic substitutions
    s = re.sub(r"\b(get)\s+po\b", r"\1 pod", s)
    s = re.sub(r"\b(get)\s+pods\b", r"\1 pod", s)

    # unify -o yaml/json
    s = re.sub(r"\s*-o\s*(yaml|json)\b", r" -o<out>", s, flags=re.IGNORECASE)
    s = re.sub(r"\s*-o(yaml|json)\b", r" -o<out>", s, flags=re.IGNORECASE)

    s = compact_ws(s)
    s = re.sub(r"\s*\|\s*", "|", s)
    return s


def command_signature(cmd: str) -> str:
    """
    similarity(semantic-ish) 계산용.
    - 숫자/버전 노이즈 감소
    """
    s = normalize_command(cmd).lower()

    s = re.sub(r"\b\d+(?:\.\d+){1,8}\b", "0.0.0", s)
    s = re.sub(r"\b\d+\b", "0", s)

    tokens = re.findall(r"[a-z0-9][a-z0-9._/:-]*|--[a-z0-9-]+|-[a-z]\b", s)

    mapped = []
    for t in tokens:
        if t == "po":
            t = "pod"
        if t in {"-oyaml", "-ojson"}:
            t = "-o<out>"
        mapped.append(t)

    kept = [t for t in mapped if t not in STOP_WORDS and len(t) > 1]
    return " ".join(kept)


def extract_keywords(s: str) -> List[str]:
    if not s:
        return []
    parts = re.split(r"[^a-zA-Z0-9]+", s.lower())
    out, seen = [], set()
    for p in parts:
        p = p.strip()
        if len(p) >= 3 and p not in seen:
            out.append(p)
            seen.add(p)
    return out


def build_guard_keywords(item_name: str, cmd_expected: str) -> List[str]:
    base = extract_keywords(item_name)
    extra = extract_keywords(cmd_expected)

    nm = (item_name or "").lower()
    if "csi-unity" in nm or "unity" in nm:
        base += ["unity"]
    if "dell-csi-operator" in nm or "dell" in nm:
        base += ["dell", "csm", "operator"]
    if "grafana_plugins_init" in nm:
        base += ["grafana", "plugins", "init"]
    if "grafana-operator" in nm:
        base += ["grafana", "operator", "grafana-operator"]
    if "kube-rbac-proxy" in nm:
        base += ["rbac", "proxy", "kube"]

    merged, seen = [], set()
    for k in base + extra:
        if k not in seen:
            merged.append(k)
            seen.add(k)
    return merged


def contains_any_keyword(hay: str, keywords: List[str]) -> bool:
    if not keywords:
        return True
    h = (hay or "").lower()
    return any(k in h for k in keywords)


def safe_cmd_for_source(cmd: str) -> str:
    return (cmd or "").replace('"', "'").strip()


def entity_tokens_from_command(cmd: str, item_name: str = "") -> List[str]:
    """
    엔티티 기반 토큰 추출:
    - -n <ns> / --namespace <ns>
    - grep <pattern>
    - path 마지막 세그먼트(예: .../grafana)
    - item_name 키워드
    """
    s = strip_quotes_unescape(str(cmd or "")).lower()

    ns = []
    for m in re.finditer(r"(?:\s|^)-(?:n)\s+([a-z0-9-]+)", s):
        ns.append(m.group(1))
    for m in re.finditer(r"(?:\s|^)--namespace\s+([a-z0-9-]+)", s):
        ns.append(m.group(1))

    greps = []
    s2 = re.sub(r"\|", " ", s)
    parts = s2.split()
    for i, tok in enumerate(parts[:-1]):
        if tok == "grep":
            greps.append(parts[i + 1])

    grep_words = []
    for g in greps:
        g = g.strip('"').strip("'")
        segs = re.split(r"[^\w./:-]+", g)
        for seg in segs:
            seg = seg.strip()
            if "/" in seg:
                last = seg.split("/")[-1]
                if len(last) >= 3:
                    grep_words.append(last)
            if len(seg) >= 3:
                grep_words.append(seg)

    path_words = []
    for m in re.finditer(r"([a-z0-9._-]+/[a-z0-9._-]+)+", s):
        p = m.group(0)
        last = p.split("/")[-1]
        if len(last) >= 3:
            path_words.append(last)

    item_kws = build_guard_keywords(item_name, cmd)

    tokens = []
    for t in ns + grep_words + path_words + item_kws:
        t = re.sub(r"[^a-z0-9._-]+", "", t)
        if not t or len(t) < 3:
            continue
        if t == "csi-unity":
            t = "unity"
        tokens.append(t)

    cleaned, seen = [], set()
    for t in tokens:
        if t in GENERIC_CMD_TOKENS:
            continue
        if t not in seen:
            cleaned.append(t)
            seen.add(t)
    return cleaned


# =========================
# Log structures
# =========================
@dataclass
class CommandBlock:
    cmd_raw: str
    cmd_norm: str
    cmd_line: int
    data_start_line: int
    data_lines: List[Tuple[int, str]]


class CommandBlockExtractor:
    def __init__(self):
        self._cache: Dict[Path, List[CommandBlock]] = {}

    def parse(self, log_path: Path) -> List[CommandBlock]:
        if log_path in self._cache:
            return self._cache[log_path]

        blocks: List[CommandBlock] = []
        current_cmd = None
        current_cmd_line = 0
        current_data: List[Tuple[int, str]] = []
        data_start_line = 0

        try:
            with log_path.open("r", encoding="utf-8", errors="ignore") as f:
                for idx, line in enumerate(f, start=1):
                    m = CMD_LINE_RE.match(line)
                    if m:
                        if current_cmd is not None:
                            blocks.append(CommandBlock(
                                cmd_raw=current_cmd,
                                cmd_norm=normalize_command(current_cmd),
                                cmd_line=current_cmd_line,
                                data_start_line=data_start_line or (current_cmd_line + 1),
                                data_lines=current_data[:],
                            ))
                        current_cmd = m.group(1).strip()
                        current_cmd_line = idx
                        current_data = []
                        data_start_line = idx + 1
                        continue

                    if current_cmd is not None:
                        if END_MARKER_RE.search(line):
                            blocks.append(CommandBlock(
                                cmd_raw=current_cmd,
                                cmd_norm=normalize_command(current_cmd),
                                cmd_line=current_cmd_line,
                                data_start_line=data_start_line,
                                data_lines=current_data[:],
                            ))
                            current_cmd = None
                            current_cmd_line = 0
                            current_data = []
                            data_start_line = 0
                        else:
                            current_data.append((idx, line.rstrip("\n")))
        except Exception:
            blocks = []

        self._cache[log_path] = blocks
        return blocks


# =========================
# LogIndexer
# =========================
class LogIndexer:
    def __init__(self, root_dir: Path):
        self.root_dir = root_dir
        self._all_logs: List[Path] = []
        self._server_cache: Dict[str, Optional[Path]] = {}
        self._scan_logs()

    def _scan_logs(self):
        self._all_logs = []
        for p in self.root_dir.rglob("*.log"):
            if p.is_file():
                self._all_logs.append(p)

    def find_log_for_server(self, server: str) -> Optional[Path]:
        if not server:
            return None
        key = server.lower()
        if key in self._server_cache:
            return self._server_cache[key]

        candidates = [p for p in self._all_logs if key in p.name.lower()]

        if not candidates:
            for p in self._all_logs:
                try:
                    found = False
                    with p.open("r", encoding="utf-8", errors="ignore") as f:
                        for line in f:
                            if key in line.lower():
                                found = True
                                break
                    if found:
                        candidates.append(p)
                except Exception:
                    continue

        if not candidates:
            self._server_cache[key] = None
            return None

        def rank(p: Path):
            try:
                mtime = p.stat().st_mtime
            except Exception:
                mtime = 0
            rel = str(p.relative_to(self.root_dir))
            return (-mtime, len(rel), rel)

        candidates.sort(key=rank)
        chosen = candidates[0]
        self._server_cache[key] = chosen
        return chosen

    def relpath(self, path: Path) -> str:
        try:
            return str(path.relative_to(self.root_dir))
        except Exception:
            return str(path)


# =========================
# ValuePicker (E-shape follow)
# =========================
@dataclass
class TokenTemplate:
    prefix: str
    suffix: str
    v_prefix: bool
    dot_count: int
    tail: str
    tail_norm: str


class ValuePicker:
    TOKEN_RE = re.compile(r"(?i)\b(v?\d+(?:\.\d+){1,8})([-+][0-9a-zA-Z._-]+)?\b")

    @staticmethod
    def _norm_tail(tail: str) -> str:
        return re.sub(r"\d+", "0", tail or "")

    @staticmethod
    def build_template_from_E(e_value: str) -> Optional[TokenTemplate]:
        if not e_value:
            return None
        s = str(e_value).strip()
        for m in ValuePicker.TOKEN_RE.finditer(s):
            core = m.group(1)
            tail = m.group(2) or ""
            prefix = s[:m.start(0)]
            suffix = s[m.end(0):]
            v_prefix = core.lower().startswith("v")
            core_num = core[1:] if v_prefix else core
            dot_count = core_num.count(".")
            return TokenTemplate(prefix, suffix, v_prefix, dot_count, tail, ValuePicker._norm_tail(tail))
        return None

    @staticmethod
    def _parse_semver_core(token: str) -> Optional[Tuple[int, ...]]:
        m = ValuePicker.TOKEN_RE.search(token or "")
        if not m:
            return None
        core = m.group(1)
        v_prefix = core.lower().startswith("v")
        core_num = core[1:] if v_prefix else core
        parts = core_num.split(".")
        try:
            return tuple(int(p) for p in parts)
        except Exception:
            return None

    @staticmethod
    def _token_shape_match(token: str, tpl: TokenTemplate) -> bool:
        m = ValuePicker.TOKEN_RE.search(token)
        if not m:
            return False
        core = m.group(1)
        tail = m.group(2) or ""
        v_prefix = core.lower().startswith("v")
        core_num = core[1:] if v_prefix else core
        dot_count = core_num.count(".")
        return (v_prefix == tpl.v_prefix) and (dot_count == tpl.dot_count) and (ValuePicker._norm_tail(tail) == tpl.tail_norm)

    @staticmethod
    def pick_value_from_block(e_value: str, block: CommandBlock) -> Tuple[Optional[str], Optional[int], str]:
        e_str = "" if e_value is None else str(e_value)
        tpl = ValuePicker.build_template_from_E(e_str)

        lines = [(ln, txt) for (ln, txt) in block.data_lines if txt.strip() != ""]
        if not lines:
            return None, None, "empty_block"

        if tpl:
            candidates: List[Tuple[Tuple[int, ...], str, int]] = []
            for ln, txt in lines:
                for m in ValuePicker.TOKEN_RE.finditer(txt):
                    token_full = m.group(0)
                    if ValuePicker._token_shape_match(token_full, tpl):
                        core = ValuePicker._parse_semver_core(token_full)
                        if core is None:
                            continue
                        candidates.append((core, token_full, ln))

            if not candidates:
                return None, None, "no_matching_format_from_E"

            candidates.sort(key=lambda x: x[0], reverse=True)
            _, token_best, ln_best = candidates[0]
            value = f"{tpl.prefix}{token_best}{tpl.suffix}".strip()
            return value, ln_best, "ok_token"

        first_ln, first_txt = lines[0]
        value = compact_ws(first_txt)
        return value, first_ln, "ok_line"


# =========================
# Matcher
# =========================
@dataclass
class MatchResult:
    ok: bool
    value: str
    cmd_line: Optional[int]
    value_line: Optional[int]
    mode: str
    score: float
    matched_cmd: Optional[str]
    reason: Optional[str]


class Matcher:
    def __init__(self, fallback_cmd_min: float = FALLBACK_CMD_MIN):
        self.fallback_cmd_min = fallback_cmd_min

    def _semantic_score(self, item_name: str, expected_cmd: str, block_cmd: str, block: CommandBlock) -> float:
        a = command_signature(expected_cmd)
        b = command_signature(block_cmd)
        struct = max(
            fuzz.token_set_ratio(a, b) / 100.0,
            fuzz.partial_ratio(a, b) / 100.0,
        )

        ent_a = entity_tokens_from_command(expected_cmd, item_name=item_name)
        ent_b = entity_tokens_from_command(block_cmd, item_name=item_name)

        out_hint = []
        for _, line in block.data_lines[:3]:
            out_hint += extract_keywords(line)
        for t in out_hint:
            if len(t) >= 3 and t not in ent_b and t not in GENERIC_CMD_TOKENS:
                ent_b.append(t)

        ent_score = 0.0
        if ent_a and ent_b:
            ent_score = fuzz.token_set_ratio(" ".join(sorted(set(ent_a))), " ".join(sorted(set(ent_b)))) / 100.0

        return max(struct, ent_score)

    def match(
        self,
        item_name: str,
        cmd_expected: str,
        e_value: str,
        blocks: List[CommandBlock],
        apply_item_guard: bool,
    ) -> MatchResult:
        if not blocks:
            return MatchResult(False, "NOT_FOUND", None, None, "none", 0.0, None, "no_command_blocks")

        exp_norm = normalize_command(cmd_expected)

        # 1) exact
        exact_candidates = [b for b in blocks if b.cmd_norm == exp_norm]
        if exact_candidates:
            b = max(exact_candidates, key=lambda x: x.cmd_line)
            val, vln, v_reason = ValuePicker.pick_value_from_block(e_value, b)
            if val is None:
                return MatchResult(False, "NOT_FOUND", b.cmd_line, None, "exact", 1.0, None, v_reason)
            return MatchResult(True, val, b.cmd_line, vln, "exact", 1.0, None, None)

        # 2) semantic-ish with strict guards
        keywords = build_guard_keywords(item_name, cmd_expected) if apply_item_guard else []
        tpl_exists = ValuePicker.build_template_from_E("" if e_value is None else str(e_value)) is not None

        best: Optional[Tuple[CommandBlock, float, str, str, int]] = None

        for b in blocks:
            score = self._semantic_score(item_name, cmd_expected, b.cmd_raw, b)
            if score < self.fallback_cmd_min:
                continue

            if apply_item_guard:
                hay = (b.cmd_raw + "\n" + "\n".join(t for _, t in b.data_lines))
                if not contains_any_keyword(hay, keywords):
                    continue

            val, vln, v_reason = ValuePicker.pick_value_from_block(e_value, b)

            if tpl_exists and v_reason != "ok_token":
                continue
            if val is None or vln is None:
                continue

            if best is None or score > best[1]:
                best = (b, score, val, b.cmd_raw, vln)

        if not best:
            return MatchResult(False, "NOT_FOUND", None, None, "semantic", 0.0, None, "no_block_passed_guards")

        b, score, val, matched_cmd, vln = best
        return MatchResult(True, val, b.cmd_line, vln, "semantic", score, matched_cmd, None)


# =========================
# Excel parsing
# =========================
@dataclass
class TableSpec:
    title: str
    header_row: int
    start_row: int
    end_row: int
    col_name: int
    col_cmd: int
    col_in: int
    col_out: int
    col_source: int
    table_type: str
    fixed_server: Optional[str]


class ExcelParser:
    def __init__(self, wb, sheet_name: str, col_overrides: Dict[str, Optional[str]]):
        self.wb = wb
        self.ws = wb[sheet_name]
        self.col_overrides = col_overrides

    @staticmethod
    def _cell_text(v: Any) -> str:
        if v is None:
            return ""
        return str(v).strip()

    def _row_header_map(self, row_idx: int) -> Dict[str, int]:
        ws = self.ws
        found: Dict[str, int] = {}
        for col in range(1, ws.max_column + 1):
            txt = self._cell_text(ws.cell(row=row_idx, column=col).value).lower()
            if not txt:
                continue
            if txt in HDR_NAME:
                found["name"] = col
            elif txt in HDR_CMD:
                found["cmd"] = col
            elif txt in HDR_IN:
                found["in"] = col
            elif txt in HDR_OUT:
                found["out"] = col
            elif txt in HDR_SOURCE:
                found["source"] = col
        return found

    def _col_from_ov(self, key: str) -> Optional[int]:
        c = self.col_overrides.get(key)
        if not c:
            return None
        try:
            return column_index_from_string(c.upper())
        except Exception:
            return None

    def _classify_table(self, title: str) -> Tuple[str, Optional[str]]:
        t = (title or "").strip()
        tl = t.lower()
        if "os version" in tl or "kernel version" in tl or "kernal version" in tl:
            return "os_kernel", None
        if tl.startswith("other version"):
            m = re.search(r"\(([^)]+)\)", t)
            if m:
                return "other_version", m.group(1).strip()
            return "other_version", None
        return "unknown", None

    def get_tables(self) -> List[TableSpec]:
        ws = self.ws
        tables: List[TableSpec] = []

        ov_name = self._col_from_ov("name")
        ov_cmd = self._col_from_ov("cmd")
        ov_in = self._col_from_ov("in")
        ov_out = self._col_from_ov("out")

        header_rows: List[Tuple[int, Dict[str, int]]] = []
        for r in range(1, ws.max_row + 1):
            m = self._row_header_map(r)
            if {"name", "cmd", "in", "out"}.issubset(m.keys()):
                header_rows.append((r, m))

        if not header_rows:
            if all([ov_name, ov_cmd, ov_in, ov_out]):
                title = self._cell_text(ws.cell(row=1, column=ov_name).value)
                return [TableSpec(
                    title=title,
                    header_row=1,
                    start_row=2,
                    end_row=ws.max_row,
                    col_name=ov_name,
                    col_cmd=ov_cmd,
                    col_in=ov_in,
                    col_out=ov_out,
                    col_source=ov_out + 1,
                    table_type="unknown",
                    fixed_server=None
                )]
            return []

        header_row_indices = [hr for hr, _ in header_rows]

        for i, (hr, m) in enumerate(header_rows):
            next_hr = header_rows[i + 1][0] if i + 1 < len(header_rows) else None

            col_name = ov_name or m["name"]
            col_cmd = ov_cmd or m["cmd"]
            col_in = ov_in or m["in"]
            col_out = ov_out or m["out"]
            col_source = col_out + 1

            title = ""
            if hr > 1:
                title = self._cell_text(ws.cell(row=hr - 1, column=col_name).value)

            start = hr + 1
            end = start - 1
            blank_name_streak = 0

            r = start
            while r <= ws.max_row:
                if next_hr and r >= next_hr:
                    break
                if r in header_row_indices:
                    break

                name_txt = self._cell_text(ws.cell(row=r, column=col_name).value)
                if name_txt == "":
                    blank_name_streak += 1
                else:
                    blank_name_streak = 0

                if blank_name_streak >= 2:
                    break

                end = r
                r += 1

            if end < start:
                continue

            table_type, fixed_server = self._classify_table(title)

            tables.append(TableSpec(
                title=title,
                header_row=hr,
                start_row=start,
                end_row=end,
                col_name=col_name,
                col_cmd=col_cmd,
                col_in=col_in,
                col_out=col_out,
                col_source=col_source,
                table_type=table_type,
                fixed_server=fixed_server,
            ))

        return tables


def insert_source_column_once(ws, tables: List[TableSpec]) -> None:
    """
    Source는 Output Data 바로 오른쪽에 새 열 1개만 삽입.
    테이블마다 반복 삽입하지 않도록 'Output+1 위치'별로 1번만 수행.
    """
    desired_positions = sorted(set(t.col_out + 1 for t in tables))

    for desired in desired_positions:
        already = False
        for t in tables:
            if t.col_out + 1 != desired:
                continue
            v = ws.cell(row=t.header_row, column=desired).value
            if str(v).strip().lower() == "source":
                already = True
                break

        if not already:
            ws.insert_cols(desired, 1)

            for t in tables:
                def bump(x: int) -> int:
                    return x + 1 if x >= desired else x

                t.col_name = bump(t.col_name)
                t.col_cmd = bump(t.col_cmd)
                t.col_in = bump(t.col_in)
                t.col_out = bump(t.col_out)
                t.col_source = t.col_out + 1

    for t in tables:
        t.col_source = t.col_out + 1
        ws.cell(row=t.header_row, column=t.col_source).value = "Source"


# =========================
# Processor
# =========================
class Processor:
    def __init__(
        self,
        excel_path: Path,
        sheet_name: str,
        log_root: Path,
        overwrite: bool,
        col_overrides: Dict[str, Optional[str]],
        q: queue.Queue,
    ):
        self.excel_path = excel_path
        self.sheet_name = sheet_name
        self.log_root = log_root
        self.overwrite = overwrite
        self.col_overrides = col_overrides
        self.q = q

        self.log_indexer = LogIndexer(log_root)
        self.extractor = CommandBlockExtractor()
        self.matcher = Matcher(FALLBACK_CMD_MIN)

    def _emit(self, kind: str, payload: Any):
        self.q.put((kind, payload))

    def run(self):
        try:
            self._emit("log", f"[{now_ts()}] Loading workbook: {self.excel_path}")
            wb = load_workbook(self.excel_path)
            parser = ExcelParser(wb, self.sheet_name, self.col_overrides)
            tables = parser.get_tables()

            if not tables:
                self._emit("error", "No tables detected. 헤더(Name/Command/Input Data/Output Data) 또는 열 강제 지정 확인.")
                return

            ws = wb[self.sheet_name]

            self._emit("log", f"[{now_ts()}] Insert Source column (once per Output position)...")
            insert_source_column_once(ws, tables)

            total_rows = sum((t.end_row - t.start_row + 1) for t in tables)
            done = 0
            self._emit("progress_max", total_rows)

            for t in tables:
                self._emit("log", f"[{now_ts()}] Table: '{t.title}' type={t.table_type} rows={t.start_row}-{t.end_row}")

                for r in range(t.start_row, t.end_row + 1):
                    name_val = ws.cell(row=r, column=t.col_name).value
                    cmd_val = ws.cell(row=r, column=t.col_cmd).value
                    in_val = ws.cell(row=r, column=t.col_in).value

                    name_txt = "" if name_val is None else str(name_val).strip()
                    cmd_txt = "" if cmd_val is None else str(cmd_val).strip()
                    e_txt = "" if in_val is None else str(in_val).strip()

                    # ✅ Command 비어있으면(제목줄/노이즈) 처리 금지
                    if cmd_txt == "":
                        done += 1
                        self._emit("progress", done)
                        continue

                    # 완전 빈 행만 스킵
                    if name_txt == "" and cmd_txt == "":
                        done += 1
                        self._emit("progress", done)
                        continue

                    if t.table_type == "other_version":
                        server = (t.fixed_server or "").strip()
                        item_name = name_txt
                        apply_item_guard = True
                    else:
                        server = name_txt
                        item_name = name_txt
                        apply_item_guard = False

                    cmd_for_src = safe_cmd_for_source(cmd_txt)

                    if not server:
                        # ✅ Output Data 셀이 비어있을 때만 입력
                        existing_out = ws.cell(row=r, column=t.col_out).value
                        if existing_out is None or str(existing_out).strip() == "":
                            ws.cell(row=r, column=t.col_out).value = "NOT_FOUND"
                            ws.cell(row=r, column=t.col_source).value = f'file=; reason=no_log_candidates; server={server}; cmd="{cmd_for_src}"'
                        done += 1
                        self._emit("progress", done)
                        continue

                    log_path = self.log_indexer.find_log_for_server(server)
                    if not log_path:
                        # ✅ Output Data 셀이 비어있을 때만 입력
                        existing_out = ws.cell(row=r, column=t.col_out).value
                        if existing_out is None or str(existing_out).strip() == "":
                            ws.cell(row=r, column=t.col_out).value = "NOT_FOUND"
                            ws.cell(row=r, column=t.col_source).value = f'file=; reason=no_log_candidates; server={server}; cmd="{cmd_for_src}"'
                        done += 1
                        self._emit("progress", done)
                        continue

                    blocks = self.extractor.parse(log_path)
                    rel = self.log_indexer.relpath(log_path)

                    if not blocks:
                        # ✅ Output Data 셀이 비어있을 때만 입력
                        existing_out = ws.cell(row=r, column=t.col_out).value
                        if existing_out is None or str(existing_out).strip() == "":
                            ws.cell(row=r, column=t.col_out).value = "NOT_FOUND"
                            ws.cell(row=r, column=t.col_source).value = f'file={rel}; reason=no_command_blocks; server={server}; cmd="{cmd_for_src}"'
                        done += 1
                        self._emit("progress", done)
                        continue

                    mr = self.matcher.match(
                        item_name=item_name,
                        cmd_expected=cmd_txt,
                        e_value=e_txt,
                        blocks=blocks,
                        apply_item_guard=apply_item_guard,
                    )

                    # ✅ Output Data 셀이 비어있을 때만 입력
                    existing_out = ws.cell(row=r, column=t.col_out).value
                    if existing_out is None or str(existing_out).strip() == "":
                        if mr.ok:
                            ws.cell(row=r, column=t.col_out).value = mr.value
                            src = f"file={rel}; cmd_line={mr.cmd_line}; value_line={mr.value_line}; mode={mr.mode}; score={mr.score:.3f}"
                            if mr.mode == "semantic" and mr.matched_cmd:
                                mc = mr.matched_cmd.replace('"', "'")
                                src += f'; matched_cmd="{mc}"'
                            if mr.score < LOW_SCORE_THRESHOLD:
                                src += "; note=LOW"
                            ws.cell(row=r, column=t.col_source).value = src
                        else:
                            ws.cell(row=r, column=t.col_out).value = "NOT_FOUND"
                            reason = mr.reason or "unknown"
                            ws.cell(row=r, column=t.col_source).value = f'file={rel}; reason={reason}; server={server}; cmd="{cmd_for_src}"'

                    done += 1
                    if done % 10 == 0:
                        self._emit("log", f"[{now_ts()}] Progress {done}/{total_rows}")
                    self._emit("progress", done)

            if self.overwrite:
                out_path = self.excel_path
            else:
                out_path = self.excel_path.with_name(f"{self.excel_path.stem}_filled.xlsx")

            self._emit("log", f"[{now_ts()}] Saving: {out_path}")
            wb.save(out_path)
            self._emit("done", str(out_path))

        except Exception as e:
            self._emit("error", f"Unhandled error: {e}")


# =========================
# GUI
# =========================
class AppGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Excel Log Auto Filler (safe match)")

        self.q: queue.Queue = queue.Queue()
        self.worker: Optional[threading.Thread] = None

        self.excel_path_var = tk.StringVar()
        self.log_root_var = tk.StringVar()
        self.sheet_var = tk.StringVar()
        self.overwrite_var = tk.BooleanVar(value=False)

        self.col_name_var = tk.StringVar()
        self.col_cmd_var = tk.StringVar()
        self.col_in_var = tk.StringVar()
        self.col_out_var = tk.StringVar()

        self._build_ui()
        self._poll_queue()

    def _build_ui(self):
        frm = ttk.Frame(self.root, padding=10)
        frm.grid(row=0, column=0, sticky="nsew")

        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        frm.columnconfigure(1, weight=1)

        ttk.Label(frm, text="Excel 파일 (.xlsx)").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.excel_path_var).grid(row=0, column=1, sticky="ew", padx=5)
        ttk.Button(frm, text="선택", command=self._choose_excel).grid(row=0, column=2, sticky="ew")

        ttk.Label(frm, text="로그 폴더(최상위)").grid(row=1, column=0, sticky="w")
        ttk.Entry(frm, textvariable=self.log_root_var).grid(row=1, column=1, sticky="ew", padx=5)
        ttk.Button(frm, text="선택", command=self._choose_log_root).grid(row=1, column=2, sticky="ew")

        ttk.Label(frm, text="시트").grid(row=2, column=0, sticky="w")
        self.sheet_combo = ttk.Combobox(frm, textvariable=self.sheet_var, values=[], state="readonly")
        self.sheet_combo.grid(row=2, column=1, sticky="ew", padx=5)

        ttk.Checkbutton(frm, text="원본 덮어쓰기(기본 OFF)", variable=self.overwrite_var).grid(row=2, column=2, sticky="w")

        box = ttk.LabelFrame(frm, text="열 강제 지정(선택) - 비워두면 헤더 자동탐지", padding=8)
        box.grid(row=3, column=0, columnspan=3, sticky="ew", pady=8)

        ttk.Label(box, text="Name").grid(row=0, column=0, sticky="w")
        ttk.Entry(box, width=6, textvariable=self.col_name_var).grid(row=0, column=1, sticky="w", padx=4)
        ttk.Label(box, text="Command").grid(row=0, column=2, sticky="w")
        ttk.Entry(box, width=6, textvariable=self.col_cmd_var).grid(row=0, column=3, sticky="w", padx=4)
        ttk.Label(box, text="Input").grid(row=0, column=4, sticky="w")
        ttk.Entry(box, width=6, textvariable=self.col_in_var).grid(row=0, column=5, sticky="w", padx=4)
        ttk.Label(box, text="Output").grid(row=0, column=6, sticky="w")
        ttk.Entry(box, width=6, textvariable=self.col_out_var).grid(row=0, column=7, sticky="w", padx=4)

        note = ttk.Label(frm, text="※ Source 열은 Output 바로 오른쪽에 '1개 열만' 삽입합니다.")
        note.grid(row=4, column=0, columnspan=3, sticky="w")

        self.run_btn = ttk.Button(frm, text="실행", command=self._start)
        self.run_btn.grid(row=5, column=0, columnspan=3, sticky="ew", pady=6)

        self.pbar = ttk.Progressbar(frm, mode="determinate")
        self.pbar.grid(row=6, column=0, columnspan=3, sticky="ew", pady=6)

        ttk.Label(frm, text="로그").grid(row=7, column=0, sticky="w")
        self.log_text = tk.Text(frm, height=16, wrap="word")
        self.log_text.grid(row=8, column=0, columnspan=3, sticky="nsew")
        frm.rowconfigure(8, weight=1)

        sb = ttk.Scrollbar(frm, orient="vertical", command=self.log_text.yview)
        sb.grid(row=8, column=3, sticky="ns")
        self.log_text.configure(yscrollcommand=sb.set)

    def _log(self, msg: str):
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")

    def _choose_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if not path:
            return
        self.excel_path_var.set(path)
        self._load_sheets(path)

    def _load_sheets(self, path: str):
        try:
            wb = load_workbook(path, read_only=True)
            sheets = wb.sheetnames
            self.sheet_combo.configure(values=sheets)
            if sheets:
                self.sheet_var.set(sheets[0])
            self._log(f"[{now_ts()}] Loaded sheets: {', '.join(sheets)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load workbook: {e}")

    def _choose_log_root(self):
        path = filedialog.askdirectory()
        if not path:
            return
        self.log_root_var.set(path)

    def _collect_col_overrides(self) -> Dict[str, Optional[str]]:
        def norm_letter(v: str) -> Optional[str]:
            v = (v or "").strip().upper()
            if not v:
                return None
            if not re.fullmatch(r"[A-Z]{1,3}", v):
                return None
            return v

        return {
            "name": norm_letter(self.col_name_var.get()),
            "cmd": norm_letter(self.col_cmd_var.get()),
            "in": norm_letter(self.col_in_var.get()),
            "out": norm_letter(self.col_out_var.get()),
        }

    def _start(self):
        if self.worker and self.worker.is_alive():
            messagebox.showwarning("Running", "이미 실행 중입니다.")
            return

        excel = self.excel_path_var.get().strip()
        log_root = self.log_root_var.get().strip()
        sheet = self.sheet_var.get().strip()

        if not excel or not os.path.isfile(excel):
            messagebox.showerror("Error", "Excel 파일을 선택하세요.")
            return
        if not log_root or not os.path.isdir(log_root):
            messagebox.showerror("Error", "로그 폴더를 선택하세요.")
            return
        if not sheet:
            messagebox.showerror("Error", "시트를 선택하세요.")
            return

        col_overrides = self._collect_col_overrides()
        self._log(f"[{now_ts()}] Start. overwrite={self.overwrite_var.get()} sheet={sheet}")
        self._log(f"[{now_ts()}] Column overrides: {col_overrides}")

        self.pbar["value"] = 0
        self.pbar["maximum"] = 1

        proc = Processor(
            excel_path=Path(excel),
            sheet_name=sheet,
            log_root=Path(log_root),
            overwrite=self.overwrite_var.get(),
            col_overrides=col_overrides,
            q=self.q
        )

        self.run_btn.configure(state="disabled")
        self.worker = threading.Thread(target=proc.run, daemon=True)
        self.worker.start()

    def _poll_queue(self):
        try:
            while True:
                kind, payload = self.q.get_nowait()
                if kind == "log":
                    self._log(str(payload))
                elif kind == "progress_max":
                    mx = int(payload) if payload else 1
                    self.pbar["maximum"] = max(mx, 1)
                elif kind == "progress":
                    self.pbar["value"] = int(payload)
                elif kind == "done":
                    out_path = str(payload)
                    self._log(f"[{now_ts()}] DONE. Output: {out_path}")
                    messagebox.showinfo("Done", f"완료!\n저장 파일:\n{out_path}")
                    self.run_btn.configure(state="normal")
                elif kind == "error":
                    self._log(f"[{now_ts()}] ERROR: {payload}")
                    messagebox.showerror("Error", str(payload))
                    self.run_btn.configure(state="normal")
        except queue.Empty:
            pass
        self.root.after(100, self._poll_queue)


def main():
    root = tk.Tk()
    root.geometry("980x640")
    AppGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
