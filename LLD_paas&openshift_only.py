import os
import re
import threading
import queue
import traceback
from dataclasses import dataclass
from typing import List, Optional, Tuple, Iterable, Any

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# rapidfuzz는 선택(없어도 동작)
try:
    from rapidfuzz import fuzz  # type: ignore
    HAS_RAPIDFUZZ = True
except Exception:
    HAS_RAPIDFUZZ = False


END_MARKER = "[END]"
END_LINE_PREFIX = "----------------------------------------------------------------------------------------------------"
DEFAULT_NOT_FOUND = "NOT_FOUND"
FORMULA_PREFIXES = ("=", "+", "-", "@")


# 요구사항 고정: Output Data는 G열(7), Old hint는 F열(6), Source는 H열(8)
COL_OLD = 6
COL_OUTPUT = 7
COL_SOURCE = 8


# -----------------------------
# 데이터 구조
# -----------------------------
@dataclass
class TableRegion:
    header_row: int
    key_row: int
    command_key: str  # 헤더 바로 위 B셀: "# <command>"
    param_col: int
    start_row: int
    end_row: int


@dataclass
class BlockParsed:
    file_path: str
    rel_path: str
    root_name: str  # openshift/paas
    cmd_line_no: int
    cmd_text: str
    # 결과를 라인 번호와 함께 보관
    result_lines: List[Tuple[int, str]]  # (line_no, line_text)
    pick: str  # exact_command / fuzzy_command / keyword_match / cosine_best
    score: Optional[float] = None


@dataclass
class ExtractResult:
    value_one_line: str
    mode: str  # full/partial
    pick: str
    score: Optional[float]
    src_line: Optional[int]  # "실제 선택된 데이터" 라인
    source: str


# -----------------------------
# 유틸: 문자열 정규화/유사도
# -----------------------------
_uuid_re = re.compile(r"\b[0-9a-fA-F]{8}\b-[0-9a-fA-F]{4}\b-[0-9a-fA-F]{4}\b-[0-9a-fA-F]{4}\b-[0-9a-fA-F]{12}\b")
_iso_ts_re = re.compile(r"\b\d{4}-\d{2}-\d{2}[T ]\d{2}:\d{2}:\d{2}(?:\.\d+)?(?:Z|[+-]\d{2}:\d{2})?\b")
_time_like_re = re.compile(r"\b\d{2}:\d{2}:\d{2}\b")
_hexlong_re = re.compile(r"\b0x[0-9a-fA-F]{8,}\b|\b[0-9a-fA-F]{16,}\b")
_bigint_re = re.compile(r"\b\d{6,}\b")
_ws_re = re.compile(r"[ \t]+")


def normalize_text_for_similarity(s: str) -> str:
    if s is None:
        return ""
    s = str(s)
    s = _uuid_re.sub(" <UUID> ", s)
    s = _iso_ts_re.sub(" <TS> ", s)
    s = _time_like_re.sub(" <TIME> ", s)
    s = _hexlong_re.sub(" <HEX> ", s)
    s = _bigint_re.sub(" <NUM> ", s)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = _ws_re.sub(" ", s)
    return s.strip()


def cosine_best_score(query: str, docs: List[str]) -> Tuple[int, float]:
    if not docs:
        return -1, 0.0
    qn = normalize_text_for_similarity(query)
    dn = [normalize_text_for_similarity(d) for d in docs]
    corpus = [qn] + dn
    vectorizer = TfidfVectorizer(analyzer="word", ngram_range=(1, 2), min_df=1)
    X = vectorizer.fit_transform(corpus)
    sims = cosine_similarity(X[0:1], X[1:])[0]
    best_i = int(sims.argmax())
    return best_i, float(sims[best_i])


def fuzzy_score(a: str, b: str) -> float:
    if not HAS_RAPIDFUZZ:
        return 0.0
    a2 = normalize_text_for_similarity(a)
    b2 = normalize_text_for_similarity(b)
    return float(fuzz.token_set_ratio(a2, b2)) / 100.0


def extract_last_key_token(param_name: str) -> Optional[str]:
    if not param_name:
        return None
    s = str(param_name).strip()
    if not s:
        return None
    parts = re.split(r"[.\[\]/\s]+", s)
    parts = [p for p in parts if p and p not in ("*",)]
    return parts[-1] if parts else None


def to_one_line_preserve_yaml(s: str) -> str:
    if s is None:
        s = ""
    s = str(s)
    s = s.replace("\t", "  ")
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\n", r"\n")
    return s


def force_excel_text(cell: Cell, value: str) -> None:
    if value is None:
        value = ""
    s = str(value)
    if s.startswith(FORMULA_PREFIXES):
        s = "'" + s
    cell.value = s
    cell.data_type = "s"


def normalize_command(s: str) -> str:
    """command 비교용: 소문자 + 공백 정규화"""
    if s is None:
        return ""
    t = str(s).strip()
    if t.startswith("#"):
        t = t[1:].strip()
    t = re.sub(r"\s+", " ", t).strip().lower()
    return t


def clean_command_key(s: str) -> str:
    if s is None:
        return ""
    t = str(s).strip()
    if t.startswith("#"):
        t = t[1:].strip()
    return t


def extract_yaml_name_from_text(s: str) -> Optional[str]:
    if not s:
        return None
    m = re.search(r"([\w.\-]+\.ya?ml)\b", str(s))
    return m.group(1) if m else None


def extract_command_keywords(cmd: str) -> List[str]:
    """
    oc/kubectl get 계열 기준키에서 후보 파일 찾기 위한 키워드 추출.
    - 예: oc get NetworkPolicy -n xxx allow-from-yyy -o yaml
      -> ["oc get", "networkpolicy", "allow-from-yyy", "xxx"]
    """
    if not cmd:
        return []
    c = normalize_command(cmd)
    toks = c.split()
    if not toks:
        return []

    keys = []
    joined = " ".join(toks)

    if "oc get" in joined:
        keys.append("oc get")
    if "kubectl get" in joined:
        keys.append("kubectl get")

    # get 다음 토큰(리소스)
    try:
        gi = toks.index("get")
        if gi + 1 < len(toks):
            keys.append(toks[gi + 1])
        # 그 다음(이름)도 시도
        if gi + 2 < len(toks) and not toks[gi + 2].startswith("-"):
            keys.append(toks[gi + 2])
    except ValueError:
        pass

    # -n namespace
    for i, t in enumerate(toks):
        if t == "-n" and i + 1 < len(toks):
            keys.append(toks[i + 1])

    blacklist = {"-o", "yaml", "json", "--", "-n"}
    keys = [k for k in keys if k and k not in blacklist]

    uniq = []
    for k in keys:
        if k not in uniq:
            uniq.append(k)
    return uniq


# -----------------------------
# 로그 탐색/파싱
# -----------------------------
def iter_log_files(root_dir: str) -> Iterable[str]:
    for base, _, files in os.walk(root_dir):
        for fn in files:
            if fn.lower().endswith(".log"):
                yield os.path.join(base, fn)


def scan_file_contains_any(fp: str, keywords: List[str], max_lines: int = 25000) -> bool:
    """스트리밍: 키워드 중 하나라도 있으면 True"""
    kws = [k for k in keywords if k]
    if not kws:
        return False
    try:
        with open(fp, "r", encoding="utf-8", errors="ignore") as f:
            for i, line in enumerate(f, start=1):
                if i > max_lines:
                    break
                for k in kws:
                    if k in line:
                        return True
        return False
    except Exception:
        return False


def parse_blocks_from_file(fp: str, rel_path: str, root_name: str) -> List[Tuple[int, str, List[Tuple[int, str]]]]:
    """
    반환:
      (cmd_line_no, cmd_text, result_lines[(line_no, text)])
    """
    blocks: List[Tuple[int, str, List[Tuple[int, str]]]] = []
    try:
        with open(fp, "r", encoding="utf-8", errors="ignore") as f:
            in_block = False
            cmd_line_no = -1
            cmd_text = ""
            buf: List[Tuple[int, str]] = []

            line_no = 0
            for raw in f:
                line_no += 1
                line = raw.rstrip("\n")

                # 실제 명령어인지 확인 (# } 같은 주석이 아닌 경우에만 새 블록)
                if line.startswith("# "):
                    cmd_part = line[2:].strip()
                    # 명령어 키워드 목록
                    CMD_KEYWORDS = (
                        'oc ', 'cat ', 'kubectl ', 'echo ', 'grep ', 'ls ', 'pwd', 'cd ',
                        'ssh ', 'scp ', 'curl ', 'wget ', 'yum ', 'dnf ', 'rpm ', 'systemctl ',
                        'docker ', 'podman ', 'crictl ', 'etcdctl ', 'openssl ', 'hostname',
                        'ip ', 'netstat ', 'ss ', 'ping ', 'traceroute ', 'nslookup ', 'dig ',
                        'df ', 'du ', 'free ', 'top ', 'ps ', 'kill ', 'pkill ', 'pgrep ',
                        'chmod ', 'chown ', 'cp ', 'mv ', 'rm ', 'mkdir ', 'touch ', 'tar ',
                        'vi ', 'vim ', 'nano ', 'head ', 'tail ', 'less ', 'more ', 'wc ',
                        'awk ', 'sed ', 'cut ', 'sort ', 'uniq ', 'find ', 'xargs ', 'tee ',
                        'date', 'uptime', 'whoami', 'id ', 'env', 'export ', 'source ', 'bash ',
                        'sh ', 'python ', 'perl ', 'ruby ', 'java ', 'node ', 'npm ', 'haproxy',
                    )
                    is_command = any(cmd_part.lower().startswith(kw) for kw in CMD_KEYWORDS)
                    
                    if is_command:
                        # 이전 블록이 END 없이 끊겼으면 저장
                        if in_block and cmd_text:
                            blocks.append((cmd_line_no, cmd_text, buf))
                        in_block = True
                        cmd_line_no = line_no
                        cmd_text = line[2:].strip()
                        buf = []
                        continue

                if in_block:
                    if END_MARKER in line and line.startswith(END_LINE_PREFIX):
                        blocks.append((cmd_line_no, cmd_text, buf))
                        in_block = False
                        cmd_line_no = -1
                        cmd_text = ""
                        buf = []
                    else:
                        buf.append((line_no, line))

            if in_block and cmd_text:
                blocks.append((cmd_line_no, cmd_text, buf))
    except Exception:
        pass
    return blocks


# -----------------------------
# YAML 후보 생성(Partial): 라인만 사용
# -----------------------------
def build_line_candidates(lines: List[Tuple[int, str]]) -> List[Tuple[str, int]]:
    out = []
    for ln, txt in lines:
        if txt.strip():
            out.append((txt, ln))
    return out


# -----------------------------
# Full/Partial 판정
# -----------------------------
def is_full_mode_hint(hint: Optional[str]) -> bool:
    """
    full 모드 판정:
    - 힌트가 비었거나(N/A 등) / 여러 줄 YAML 덩어리일 때만 full
    - 'metadata:' 같은 단일 키는 절대 full로 보지 않는다
    """
    if hint is None:
        return True
    s = str(hint).strip()
    if not s:
        return True
    if s.lower() in ("n/a", "na", "none", "null", "-", "--"):
        return True

    # ✅ 여러 줄이면: full로 취급(기존 의도 유지)
    if "\n" in s:
        return True

    # ✅ 단일 키/짧은 문자열은 full로 취급하지 않는다
    # (기존 코드의 spec:/metadata:/apiVersion:/kind: 예외처리가 문제였음)
    return False


# -----------------------------
# 시트 처리
# -----------------------------
class BaseSheetHandler:
    def __init__(self, sheet: Worksheet, log_root: str, sub_root: str, ui_logger):
        self.sheet = sheet
        self.log_root = log_root
        self.sub_root = sub_root
        self.search_root = os.path.join(log_root, sub_root)
        self.ui_logger = ui_logger

    def log(self, msg: Any):
        self.ui_logger(("LOG", str(msg)))

    # ✅ 시트당 1회만 H열 보장(중요)
    def ensure_source_column_once(self):
        ws = self.sheet
        # 이미 H열 헤더가 Output Source면 그대로
        for r in range(1, min(ws.max_row, 80) + 1):
            v = ws.cell(r, COL_SOURCE).value
            if isinstance(v, str) and v.strip().lower() == "output source":
                return

        # H열(8)이 비어있거나 다른 값이면, "G 오른쪽에 삽입"이 원칙
        # 즉, 항상 COL_SOURCE 위치에 insert_cols(8,1)
        ws.insert_cols(COL_SOURCE, 1)
        # 헤더는 테이블마다 세팅하지만, 최소한 1행에도 써둔다(안전)
        ws.cell(1, COL_SOURCE).value = "Output Source"

    def detect_tables(self) -> List[TableRegion]:
        ws = self.sheet
        max_row = ws.max_row
        max_col = ws.max_column

        def cell_text(r, c) -> str:
            v = ws.cell(r, c).value
            return "" if v is None else str(v)

        tables: List[TableRegion] = []
        param_keys = ("parameter name", "parameter", "param name")
        output_keys = ("output data", "output", "result")

        for r in range(1, max_row + 1):
            row_texts = [cell_text(r, c) for c in range(1, max_col + 1)]
            lowered = [t.strip().lower() for t in row_texts]

            has_param = any(any(k in t for k in param_keys) for t in lowered)
            has_output = any(any(k in t for k in output_keys) for t in lowered)
            if not (has_param and has_output):
                continue

            # param 컬럼 찾기(정확히)
            param_col = -1
            for c in range(1, max_col + 1):
                t = cell_text(r, c).strip().lower()
                if any(k in t for k in param_keys):
                    param_col = c
                    break
            if param_col < 0:
                continue

            key_row = r - 1
            if key_row < 1:
                continue

            command_key = ws.cell(key_row, 2).value  # B
            if command_key is None:
                continue
            command_key = str(command_key).strip()
            if not command_key.startswith("#"):
                continue

            start_row = r + 1

            blank_streak = 0
            end_row = start_row - 1
            rr = start_row
            while rr <= max_row:
                # 다음 헤더 만나면 종료
                row_t = [cell_text(rr, cc).strip().lower() for cc in range(1, max_col + 1)]
                if any(any(k in x for k in param_keys) for x in row_t) and any(any(k in x for k in output_keys) for x in row_t):
                    break

                # ✅ B열에 새 명령어("# ")가 나오면 테이블 종료
                b_val = cell_text(rr, 2).strip()
                if b_val.startswith("#"):
                    break

                pv = ws.cell(rr, param_col).value
                if pv is None or str(pv).strip() == "":
                    blank_streak += 1
                else:
                    blank_streak = 0
                    end_row = rr

                # ✅ 빈 행이 1개만 있어도 테이블 종료 (더 엄격하게)
                if blank_streak >= 1:
                    break
                rr += 1

            if end_row >= start_row:
                tables.append(TableRegion(
                    header_row=r,
                    key_row=key_row,
                    command_key=command_key,
                    param_col=param_col,
                    start_row=start_row,
                    end_row=end_row
                ))

        return tables

    def find_best_block_for_table(self, table: TableRegion) -> Optional[BlockParsed]:
        if not os.path.isdir(self.search_root):
            self.log(f"[ERROR] 로그 폴더 없음: {self.search_root}")
            return None

        ws = self.sheet
        command_raw = table.command_key
        command_clean = clean_command_key(command_raw)
        cmd_norm = normalize_command(command_raw)

        yaml_name = extract_yaml_name_from_text(command_raw)
        if not yaml_name:
            f_header = ws.cell(table.header_row, COL_OLD).value
            yaml_name = extract_yaml_name_from_text(f_header)

        # 후보 파일 선별 키워드
        exact_marker = "# " + command_clean  # 원문 형태 근접
        cmd_keywords = extract_command_keywords(command_clean)
        keywords = [exact_marker]
        if yaml_name:
            keywords += [yaml_name, "# cat", "cat " + yaml_name]
        keywords += cmd_keywords

        candidates: List[Tuple[str, float]] = []
        for fp in iter_log_files(self.search_root):
            try:
                mtime = os.path.getmtime(fp)
            except Exception:
                mtime = 0.0
            if scan_file_contains_any(fp, keywords):
                candidates.append((fp, mtime))

        # ✅ 후보가 하나도 없으면: 최신 로그 파일로 fallback(중요)
        if not candidates:
            all_files = []
            for fp in iter_log_files(self.search_root):
                try:
                    mtime = os.path.getmtime(fp)
                except Exception:
                    mtime = 0.0
                all_files.append((fp, mtime))
            all_files.sort(key=lambda x: x[1], reverse=True)
            candidates = all_files[:20]  # 최신 20개만 파싱해서 블록을 찾음

        if not candidates:
            return None

        candidates.sort(key=lambda x: x[1], reverse=True)
        top_files = [fp for fp, _ in candidates[:20]]

        # 테이블 힌트(F열 값 모음) - 블록 다중일 때 cosine 선택용
        hints = []
        for r in range(table.start_row, table.end_row + 1):
            hv = ws.cell(r, COL_OLD).value
            if hv is None:
                continue
            s = str(hv).strip()
            if s and s.lower() not in ("n/a", "na", "none", "null"):
                hints.append(s)
        table_hint_text = "\n".join(hints[:40]) if hints else command_raw

        best: Optional[BlockParsed] = None
        best_score = -1.0

        for fp in top_files:
            rel = os.path.relpath(fp, self.log_root).replace("\\", "/")
            blocks = parse_blocks_from_file(fp, rel, self.sub_root)
            if not blocks:
                continue

            exact_list = []
            fuzzy_list = []
            yaml_list = []

            for (cmd_ln, cmd_txt, res_lines) in blocks:
                b_norm = normalize_command(cmd_txt)
                if b_norm == cmd_norm:
                    exact_list.append((cmd_ln, cmd_txt, res_lines))
                else:
                    if cmd_norm and (cmd_norm in b_norm or b_norm in cmd_norm):
                        fuzzy_list.append((cmd_ln, cmd_txt, res_lines))

                if yaml_name:
                    if yaml_name.lower() in cmd_txt.lower():
                        if ("cat " in cmd_txt.lower()) or (" oc " in cmd_txt.lower()) or ("kubectl" in cmd_txt.lower()):
                            yaml_list.append((cmd_ln, cmd_txt, res_lines))

            def pick_best_from(list_blocks, pick_label: str) -> Optional[BlockParsed]:
                nonlocal best, best_score
                if not list_blocks:
                    return None
                docs = ["\n".join([t for _, t in res_lines]).strip("\n") for (_, _, res_lines) in list_blocks]
                bi, sc = cosine_best_score(table_hint_text, docs)
                cmd_ln, cmd_txt, res_lines = list_blocks[bi]
                bp = BlockParsed(
                    file_path=fp,
                    rel_path=rel,
                    root_name=self.sub_root,
                    cmd_line_no=cmd_ln,
                    cmd_text=cmd_txt,
                    result_lines=res_lines,
                    pick=pick_label if len(list_blocks) == 1 else "cosine_best",
                    score=sc
                )
                if sc > best_score:
                    best_score = sc
                    best = bp
                return bp

            pick_best_from(exact_list, "exact_command")
            if not exact_list:
                pick_best_from(fuzzy_list, "fuzzy_command")
            if not exact_list and not fuzzy_list:
                pick_best_from(yaml_list, "yaml_name_match")

        if best:
            self.log(f"[BLOCK] pick={best.pick}, score={best.score:.3f} file={best.rel_path} cmd_line={best.cmd_line_no}")
        return best

    def process(self) -> Tuple[int, int]:
        ws = self.sheet

        # ✅ 시트당 1회만 H열 삽입/보장(가장 중요)
        self.ensure_source_column_once()

        tables = self.detect_tables()
        if not tables:
            self.log(f"[WARN] 테이블 없음: sheet={ws.title}")
            return 0, 0

        total_rows = sum((t.end_row - t.start_row + 1) for t in tables)
        processed = 0
        not_found = 0

        for ti, table in enumerate(tables, start=1):
            ws.cell(table.header_row, COL_SOURCE).value = "Output Source"
            self.log(f"[TABLE {ti}/{len(tables)}] header_row={table.header_row}, key={table.command_key}")

            block = self.find_best_block_for_table(table)
            if block is None:
                self.log(f"[TABLE {ti}] 블록 없음 -> NOT_FOUND")
                for r in range(table.start_row, table.end_row + 1):
                    self._write_not_found(r, reason="no_block", root=self.sub_root)
                    processed += 1
                    not_found += 1
                    self.ui_logger(("PROGRESS", processed, total_rows))
                continue

            # ✅ 순서 기반 매칭: 이전 매칭 위치 이후부터만 검색
            block_lines = block.result_lines  # [(line_no, text), ...]
            last_matched_idx = -1  # 마지막으로 매칭된 인덱스

            def normalize_for_match(s: str) -> str:
                """매칭용 정규화: 앞뒤 공백 제거, 여러 공백을 하나로"""
                s = str(s).strip()
                s = re.sub(r'\s+', ' ', s)
                return s

            for r in range(table.start_row, table.end_row + 1):
                pn = ws.cell(r, table.param_col).value  # Parameter Name
                if pn is None or str(pn).strip() == "":
                    processed += 1
                    self.ui_logger(("PROGRESS", processed, total_rows))
                    continue

                # ✅ 매칭 기준: F열(hint) 값 우선, 없으면 Parameter Name
                hint = ws.cell(r, COL_OLD).value
                if hint and str(hint).strip():
                    search_text = str(hint).strip()
                else:
                    search_text = str(pn).strip()

                search_normalized = normalize_for_match(search_text)

                try:
                    # 마지막 매칭 위치 이후부터 검색
                    found_idx = -1
                    found_line_no = None
                    found_line_text = None

                    for idx in range(last_matched_idx + 1, len(block_lines)):
                        log_line_no, log_line_text = block_lines[idx]
                        
                        # 빈 라인, 주석 라인 스킵
                        stripped = log_line_text.strip()
                        if not stripped or stripped.startswith('#'):
                            continue

                        log_normalized = normalize_for_match(log_line_text)

                        # 1) 정규화 후 완전 일치
                        if log_normalized == search_normalized:
                            found_idx = idx
                            found_line_no = log_line_no
                            found_line_text = log_line_text
                            break

                        # 2) 정규화 후 시작 부분 일치
                        if log_normalized.startswith(search_normalized):
                            found_idx = idx
                            found_line_no = log_line_no
                            found_line_text = log_line_text
                            break

                        # 3) search가 log를 포함 (search가 더 길 때)
                        if search_normalized.startswith(log_normalized) and len(log_normalized) > 3:
                            found_idx = idx
                            found_line_no = log_line_no
                            found_line_text = log_line_text
                            break

                        # 4) YAML 키 비교 (콜론 앞부분)
                        if ':' in search_normalized and ':' in log_normalized:
                            search_key = search_normalized.split(':')[0].strip()
                            log_key = log_normalized.split(':')[0].strip()
                            if search_key == log_key and len(search_key) > 2:
                                found_idx = idx
                                found_line_no = log_line_no
                                found_line_text = log_line_text
                                break

                        # 5) 로그 라인이 search 텍스트를 포함
                        if search_normalized in log_normalized:
                            found_idx = idx
                            found_line_no = log_line_no
                            found_line_text = log_line_text
                            break

                    if found_idx >= 0:
                        last_matched_idx = found_idx

                        # ✅ Output Data 셀이 비어있을 때만 입력
                        existing_output = ws.cell(r, COL_OUTPUT).value
                        if existing_output is None or str(existing_output).strip() == "":
                            one = to_one_line_preserve_yaml(found_line_text)
                            source = (
                                f"root={block.root_name}; file={block.rel_path}; data_line={found_line_no}; "
                                f"mode=matched; pick=hint_match; cmd_line={block.cmd_line_no}"
                            )

                            force_excel_text(ws.cell(r, COL_OUTPUT), one)
                            force_excel_text(ws.cell(r, COL_SOURCE), source)

                            self.log(f"  [ROW {r}] mode=matched line={found_line_no}")
                        else:
                            self.log(f"  [ROW {r}] SKIP - already has value")
                    else:
                        # 매칭 실패 -> fallback: 다음 라인 순차 사용
                        fallback_idx = last_matched_idx + 1
                        while fallback_idx < len(block_lines):
                            log_line_no, log_line_text = block_lines[fallback_idx]
                            stripped = log_line_text.strip()
                            if stripped and not stripped.startswith('#'):
                                break
                            fallback_idx += 1

                        if fallback_idx < len(block_lines):
                            log_line_no, log_line_text = block_lines[fallback_idx]
                            last_matched_idx = fallback_idx

                            # ✅ Output Data 셀이 비어있을 때만 입력
                            existing_output = ws.cell(r, COL_OUTPUT).value
                            if existing_output is None or str(existing_output).strip() == "":
                                one = to_one_line_preserve_yaml(log_line_text)
                                source = (
                                    f"root={block.root_name}; file={block.rel_path}; data_line={log_line_no}; "
                                    f"mode=sequential; pick=fallback; cmd_line={block.cmd_line_no}"
                                )

                                force_excel_text(ws.cell(r, COL_OUTPUT), one)
                                force_excel_text(ws.cell(r, COL_SOURCE), source)

                                self.log(f"  [ROW {r}] mode=fallback line={log_line_no}")
                            else:
                                self.log(f"  [ROW {r}] SKIP - already has value")
                        else:
                            self._write_not_found(r, reason="no_more_lines", root=self.sub_root)
                            not_found += 1

                except Exception as e:
                    self._write_not_found(r, reason=f"exception:{e}", root=self.sub_root)
                    not_found += 1

                processed += 1
                self.ui_logger(("PROGRESS", processed, total_rows))

        return (processed - not_found), not_found

    def extract_full(self, block: BlockParsed) -> ExtractResult:
        text = "\n".join([t for _, t in block.result_lines]).strip("\n")
        one = to_one_line_preserve_yaml(text)

        src_line = block.result_lines[0][0] if block.result_lines else block.cmd_line_no

        source = (
            f"root={block.root_name}; file={block.rel_path}; data_line={src_line}; "
            f"mode=full; pick={block.pick}; score={block.score if block.score is not None else ''}; "
            f"cmd_line={block.cmd_line_no}"
        )
        return ExtractResult(one, "full", "full_block", block.score, src_line, source)

    def extract_partial(self, block: BlockParsed, param_name: Any, hint: Any) -> ExtractResult:
        hint_s = "" if hint is None else str(hint).strip()
        param_s = "" if param_name is None else str(param_name).strip()

        # ✅ partial은 "라인"만 대상으로 한다 (chunk 사용 금지)
        line_cands = build_line_candidates(block.result_lines)  # (text, line_no)
        if not line_cands:
            return self._not_found_result(block, reason="no_line_candidates")

        # query는 hint 우선
        q = hint_s if hint_s else param_s
        q_norm = normalize_text_for_similarity(q).lower()

        # -------------------------
        # 0) 우선: "정확한 라인 매칭" (가장 안전)
        # -------------------------
        target = param_s.strip()
        if target:
            # 0-1) 완전 동일(공백만 무시)
            for txt, ln in line_cands:
                if txt.strip() == target:
                    one = to_one_line_preserve_yaml(txt.strip())
                    source = (
                        f"root={block.root_name}; file={block.rel_path}; data_line={ln}; "
                        f"mode=partial; pick=exact_line; score=1.000; cmd_line={block.cmd_line_no}"
                    )
                    return ExtractResult(one, "partial", "exact_line", 1.0, ln, source)

            # 0-2) YAML 키 매칭 (예: 'metadata:'면 'metadata:'로 시작하는 라인)
            if target.endswith(":"):
                key = target[:-1].strip()
                if key:
                    pattern = re.compile(rf"^\s*{re.escape(key)}\s*:\s*$")
                    for txt, ln in line_cands:
                        if pattern.match(txt):
                            one = to_one_line_preserve_yaml(txt.rstrip())
                            source = (
                                f"root={block.root_name}; file={block.rel_path}; data_line={ln}; "
                                f"mode=partial; pick=key_line; score=1.000; cmd_line={block.cmd_line_no}"
                            )
                            return ExtractResult(one, "partial", "key_line", 1.0, ln, source)

        # -------------------------
        # 1) 코사인 유사도(라인만) + 간단 boost
        # -------------------------
        docs = [t for (t, _) in line_cands]
        best_i, best_cos = cosine_best_score(q, docs)
        best_text, best_line = line_cands[best_i]
        best_score = best_cos
        best_pick = "line_cosine"

        # 흔한 헤더 라인(apiVersion/kind/metadata 등)이 "무관한 행"에 끼어드는 것 방지
        def header_penalty(txt: str, qn: str) -> float:
            t = txt.lstrip().lower()
            # query에 해당 키가 없는데 헤더면 강하게 패널티
            if t.startswith("apiversion:") and "apiversion" not in qn:
                return -0.40
            if t.startswith("kind:") and "kind" not in qn:
                return -0.35
            if t.startswith("metadata:") and "metadata" not in qn:
                return -0.35
            if t.startswith("items:") and "items" not in qn:
                return -0.35
            return 0.0

        # 상위 80개 재랭크
        qn = normalize_text_for_similarity(q)
        dn = [normalize_text_for_similarity(d) for d in docs]
        vectorizer = TfidfVectorizer(analyzer="word", ngram_range=(1, 2), min_df=1)
        X = vectorizer.fit_transform([qn] + dn)
        sims = cosine_similarity(X[0:1], X[1:])[0]
        top_idx = sims.argsort()[::-1][:80]

        last_token = extract_last_key_token(param_s)

        for idx in top_idx:
            idx = int(idx)
            cand_text, cand_line = line_cands[idx]
            sc = float(sims[idx])

            sc += header_penalty(cand_text, q_norm)

            if last_token and (last_token.lower() + ":") in cand_text.lower():
                sc += 0.10
            elif last_token and last_token.lower() in cand_text.lower():
                sc += 0.05

            if HAS_RAPIDFUZZ:
                sc += 0.12 * fuzzy_score(q, cand_text)

            if sc > best_score:
                best_score = sc
                best_text = cand_text
                best_line = cand_line
                best_pick = "line_cosine+boost"

        one = to_one_line_preserve_yaml(best_text)

        # ✅ partial인데 너무 길면: 그냥 NOT_FOUND 처리(여기서 덩어리 나오면 무조건 이상 케이스)
        if len(one) > 500:
            return self._not_found_result(block, reason="partial_too_long_guard")

        source = (
            f"root={block.root_name}; file={block.rel_path}; data_line={best_line}; "
            f"mode=partial; pick={best_pick}; score={best_score:.3f}; cmd_line={block.cmd_line_no}"
        )
        return ExtractResult(one, "partial", best_pick, best_score, best_line, source)

    def _not_found_result(self, block: BlockParsed, reason: str) -> ExtractResult:
        source = (
            f"root={block.root_name}; file={block.rel_path}; "
            f"mode=?; pick=NOT_FOUND; reason={reason}; cmd_line={block.cmd_line_no}"
        )
        return ExtractResult(DEFAULT_NOT_FOUND, "?", "NOT_FOUND", None, None, source)

    def _write_not_found(self, row: int, reason: str, root: str):
        ws = self.sheet
        # ✅ Output Data 셀이 비어있을 때만 입력
        existing_output = ws.cell(row, COL_OUTPUT).value
        if existing_output is None or str(existing_output).strip() == "":
            force_excel_text(ws.cell(row, COL_OUTPUT), DEFAULT_NOT_FOUND)
            force_excel_text(ws.cell(row, COL_SOURCE), f"root={root}; pick=NOT_FOUND; reason={reason}")
        else:
            self.log(f"  [ROW {row}] SKIP NOT_FOUND - already has value")

    def run(self) -> Tuple[int, int]:
        return self.process()


class OpenShiftConfigHandler(BaseSheetHandler):
    def __init__(self, sheet: Worksheet, log_root: str, ui_logger):
        super().__init__(sheet, log_root, "openshift", ui_logger)


class PaasConfigHandler(BaseSheetHandler):
    def __init__(self, sheet: Worksheet, log_root: str, ui_logger):
        super().__init__(sheet, log_root, "paas", ui_logger)


def handler_for_sheet(sheet_name: str):
    s = (sheet_name or "").strip().lower()
    if "openshift" in s and "config" in s:
        return OpenShiftConfigHandler
    if "paas" in s and "config" in s:
        return PaasConfigHandler
    return None


# -----------------------------
# GUI + Worker Thread
# -----------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Log Auto Filler (G/H columns)")
        self.geometry("920x640")

        self.excel_path = tk.StringVar()
        self.log_root = tk.StringVar()
        self.sheet_name = tk.StringVar()
        self.overwrite = tk.BooleanVar(value=False)

        # ✅ 로그/진행용 큐
        self.msg_queue: "queue.Queue" = queue.Queue()
        # ✅ SaveAs 응답 전용 큐(핵심 수정)
        self.saveas_queue: "queue.Queue[str]" = queue.Queue()

        self.worker: Optional[threading.Thread] = None

        self._build_ui()
        self._poll_queue()

    def _build_ui(self):
        pad = {"padx": 8, "pady": 6}
        frm = ttk.Frame(self)
        frm.pack(fill="x", **pad)

        row1 = ttk.Frame(frm)
        row1.pack(fill="x", **pad)
        ttk.Label(row1, text="Excel (.xlsx):", width=16).pack(side="left")
        ttk.Entry(row1, textvariable=self.excel_path).pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(row1, text="파일 선택", command=self.pick_excel).pack(side="left")

        row2 = ttk.Frame(frm)
        row2.pack(fill="x", **pad)
        ttk.Label(row2, text="Log Root Folder:", width=16).pack(side="left")
        ttk.Entry(row2, textvariable=self.log_root).pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(row2, text="폴더 선택", command=self.pick_log_root).pack(side="left")

        row3 = ttk.Frame(frm)
        row3.pack(fill="x", **pad)
        ttk.Label(row3, text="Sheet:", width=16).pack(side="left")
        self.sheet_combo = ttk.Combobox(row3, textvariable=self.sheet_name, state="readonly", values=[])
        self.sheet_combo.pack(side="left", fill="x", expand=True, padx=6)
        ttk.Button(row3, text="시트 불러오기", command=self.load_sheets).pack(side="left")

        row4 = ttk.Frame(frm)
        row4.pack(fill="x", **pad)
        ttk.Checkbutton(row4, text="원본 덮어쓰기(기본 OFF)", variable=self.overwrite).pack(side="left")
        ttk.Button(row4, text="실행", command=self.run_clicked).pack(side="right")

        row5 = ttk.Frame(frm)
        row5.pack(fill="x", **pad)
        self.pbar = ttk.Progressbar(row5, orient="horizontal", mode="determinate")
        self.pbar.pack(fill="x", expand=True)

        logfrm = ttk.Frame(self)
        logfrm.pack(fill="both", expand=True, **pad)
        ttk.Label(logfrm, text="Log:").pack(anchor="w")
        self.log_text = tk.Text(logfrm, height=18, wrap="word")
        self.log_text.pack(fill="both", expand=True)
        self.log_text.configure(state="disabled")

    def pick_excel(self):
        path = filedialog.askopenfilename(title="Select Excel file", filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.excel_path.set(path)

    def pick_log_root(self):
        path = filedialog.askdirectory(title="Select log root folder")
        if path:
            self.log_root.set(path)

    def load_sheets(self):
        path = self.excel_path.get().strip()
        if not path or not os.path.isfile(path):
            messagebox.showerror("Error", "유효한 Excel 파일(.xlsx)을 선택하세요.")
            return
        try:
            wb = load_workbook(path)
            sheets = wb.sheetnames
            wb.close()
            self.sheet_combo["values"] = sheets
            if sheets:
                self.sheet_name.set(sheets[0])
            self._ui_log(f"[INFO] 시트 로드 완료: {len(sheets)}개")
        except Exception as e:
            messagebox.showerror("Error", f"시트 로드 실패: {e}")

    def run_clicked(self):
        if self.worker and self.worker.is_alive():
            messagebox.showwarning("Running", "이미 실행 중입니다.")
            return

        xlsx = self.excel_path.get().strip()
        lroot = self.log_root.get().strip()
        sname = self.sheet_name.get().strip()

        if not xlsx or not os.path.isfile(xlsx):
            messagebox.showerror("Error", "유효한 Excel 파일(.xlsx)을 선택하세요.")
            return
        if not lroot or not os.path.isdir(lroot):
            messagebox.showerror("Error", "유효한 로그 최상위 폴더를 선택하세요.")
            return
        if not sname:
            messagebox.showerror("Error", "시트를 선택하세요.")
            return

        hcls = handler_for_sheet(sname)
        if hcls is None:
            messagebox.showerror(
                "Error",
                "시트명이 매핑 규칙과 맞지 않습니다.\n- openshift config\n- paas config"
            )
            return

        self._ui_log("[INFO] 작업 시작")
        self.pbar["value"] = 0
        self.pbar["maximum"] = 1

        self.worker = threading.Thread(
            target=self._worker_main,
            args=(xlsx, lroot, sname, hcls, self.overwrite.get()),
            daemon=True
        )
        self.worker.start()

    def _worker_main(self, xlsx: str, log_root: str, sheet_name: str, handler_cls, overwrite: bool):
        wb = None
        try:
            wb = load_workbook(xlsx)
            if sheet_name not in wb.sheetnames:
                raise RuntimeError(f"Sheet not found: {sheet_name}")
            ws = wb[sheet_name]

            def ui_logger(payload):
                self.msg_queue.put(payload)

            handler = handler_cls(ws, log_root, ui_logger)
            ok, nf = handler.run()

            save_path = xlsx
            if not overwrite:
                base, ext = os.path.splitext(xlsx)
                default = base + "_filled" + ext
                self.msg_queue.put(("ASK_SAVEAS", default))

                # ✅ SaveAs 응답은 saveas_queue로만 받는다 (핵심 수정)
                save_path = self.saveas_queue.get()
                if not save_path:
                    self.msg_queue.put(("LOG", "[INFO] 저장 취소 (파일 저장 안 함)"))
                    return
                if not save_path.lower().endswith(".xlsx"):
                    save_path += ".xlsx"

            wb.save(save_path)
            self.msg_queue.put(("DONE", ok, nf, save_path))

        except Exception as e:
            tb = traceback.format_exc()
            self.msg_queue.put(("ERROR", f"{e}\n\n{tb}"))
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    def _poll_queue(self):
        try:
            while True:
                msg = self.msg_queue.get_nowait()
                self._handle_msg(msg)
        except queue.Empty:
            pass
        self.after(120, self._poll_queue)

    def _handle_msg(self, msg):
        if isinstance(msg, tuple):
            kind = msg[0]
            if kind == "LOG":
                self._ui_log(msg[1])
            elif kind == "PROGRESS":
                _, cur, total = msg
                total = max(1, int(total))
                self.pbar["maximum"] = total
                self.pbar["value"] = int(cur)
            elif kind == "ASK_SAVEAS":
                _, default_path = msg
                path = filedialog.asksaveasfilename(
                    title="Save as",
                    defaultextension=".xlsx",
                    initialfile=os.path.basename(default_path),
                    filetypes=[("Excel files", "*.xlsx")],
                )
                # ✅ 결과는 saveas_queue로만 전달 (핵심 수정)
                self.saveas_queue.put(path)
            elif kind == "DONE":
                _, ok, nf, save_path = msg
                self._ui_log(f"[DONE] success={ok}, not_found={nf}")
                self._ui_log(f"[SAVED] {save_path}")
                messagebox.showinfo("Done", f"완료!\nsuccess={ok}\nnot_found={nf}\nsaved={save_path}")
            elif kind == "ERROR":
                _, detail = msg
                self._ui_log("[ERROR]\n" + detail)
                messagebox.showerror("Error", detail)
        else:
            self._ui_log(str(msg))

    def _ui_log(self, s: str):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", str(s) + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")


def main():
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
