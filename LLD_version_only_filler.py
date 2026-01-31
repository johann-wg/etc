import os
import re
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path
from abc import ABC, abstractmethod
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from rapidfuzz import process, fuzz

# ==========================================
# 1. Base Handler & Version Implementation
# ==========================================

class BaseSheetHandler(ABC):
    """시트별 처리 로직의 인터페이스를 정의하는 클래스"""
    @abstractmethod
    def process(self, sheet, log_root, options, log_callback):
        pass

class VersionSheetHandler(BaseSheetHandler):
    """'version' 시트의 특화된 로직을 처리하는 클래스"""
    
    def __init__(self):
        self.output_source_col_idx = None

    def find_log_folder(self, root_path, sheet_name, match_mode):
        """시트명과 유사한 로그 하위 폴더 탐색"""
        folders = [d.name for d in root_path.iterdir() if d.is_dir()]
        if not folders: return None
        
        if match_mode == "포함 매칭":
            for f in folders:
                if sheet_name.lower().replace(" ", "") in f.lower().replace(" ", ""):
                    return root_path / f
        else: # 유사도 매칭
            best_match = process.extractOne(sheet_name, folders, scorer=fuzz.WRatio)
            if best_match and best_match[1] > 60:
                return root_path / best_match[0]
        return None

    def get_target_server(self, row_idx, sheet, name_col, title_server):
        """테이블 제목 또는 Name 열에서 대상 서버 결정"""
        name_val = str(sheet.cell(row=row_idx, column=name_col).value or "").strip()
        return title_server if title_server else name_val

    def find_log_file(self, folder_path, server_name):
        """서버명에 맞는 최적의 로그 파일 탐색"""
        files = list(folder_path.glob("**/*"))
        candidates = []
        for f in files:
            if f.is_file() and server_name.lower() in f.name.lower():
                candidates.append(f)
        
        if not candidates:
            # 파일 내용까지 확인하는 로직 (성능을 위해 스트리밍)
            for f in files:
                if f.is_file():
                    try:
                        with open(f, 'r', encoding='utf-8', errors='ignore') as temp_f:
                            for _ in range(20): # 상위 20줄만 확인
                                if server_name in temp_f.readline():
                                    candidates.append(f)
                                    break
                    except: continue

        if not candidates: return None
        # 수정시간 최신순, 경로 짧은순 정렬
        candidates.sort(key=lambda x: (x.stat().st_mtime, -len(str(x))), reverse=True)
        return candidates[0]

    def extract_from_log(self, file_path, command, strategy, input_data):
        """로그 파일에서 명령 블록 추출"""
        results = []
        target_cmd = f"# {command.strip()}"
        found_block = False
        line_num = 0
        cmd_line_num = 0

        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for idx, line in enumerate(f, 1):
                    clean_line = line.strip()
                    if clean_line.startswith(target_cmd):
                        found_block = True
                        cmd_line_num = idx
                        results = []
                        continue
                    
                    if found_block:
                        if "---[END]" in clean_line:
                            break
                        if clean_line:
                            results.append((idx, clean_line))
            
            if not found_block:
                return None, f"NOT_FOUND; cmd_not_found: {command}", 0
            
            if not results:
                return None, f"NOT_FOUND; empty_block", cmd_line_num

            # 추출 전략 적용
            if strategy == "첫 줄":
                res_idx, res_val = results[0]
                pick_info = "first-nonempty"
            elif strategy == "마지막 줄":
                res_idx, res_val = results[-1]
                pick_info = "last-nonempty"
            else: # 유사도 매칭
                vals = [r[1] for r in results]
                best = process.extractOne(str(input_data), vals)
                res_val = best[0]
                res_idx = results[vals.index(res_val)][0]
                pick_info = f"similarity:{best[1]:.2f}"

            return res_val, f"line={res_idx}; pick={pick_info}", cmd_line_num

        except Exception as e:
            return None, f"ERROR: {str(e)}", 0

    def process(self, sheet, log_root, options, log_callback):
        # 1. 로그 폴더 매칭
        target_folder = self.find_log_folder(Path(log_root), sheet.title, options['match_mode'])
        if not target_folder:
            log_callback(f"❌ {sheet.title}: 매칭되는 폴더를 찾을 수 없습니다.")
            return

        log_callback(f"📂 {sheet.title} 시트 처리 시작 (폴더: {target_folder.name})")
        
        # 컬럼 인덱스 설정
        name_col = column_index_from_string(options['col_name'])
        cmd_col = column_index_from_string(options['col_cmd'])
        input_col = column_index_from_string(options['col_input'])
        out_col = column_index_from_string(options['col_out'])
        
        # Output Source 열 삽입
        sheet.insert_cols(out_col + 1)
        sheet.cell(row=1, column=out_col + 1).value = "Output Source"
        
        current_title_server = None
        
        for row in range(1, sheet.max_row + 1):
            cell_val = str(sheet.cell(row=row, column=name_col).value or "").strip()
            
            # 테이블 제목 탐지 (예: Other Version (bastion01))
            title_match = re.search(r'\(([^)]+)\)', cell_val)
            if title_match and ("Version" in cell_val or "Table" in cell_val):
                current_title_server = title_match.group(1)
                continue
            
            # 헤더 행 스킵 및 데이터 행 판별 (Command가 있는 경우)
            cmd_val = sheet.cell(row=row, column=cmd_col).value
            if not cmd_val or cmd_val == "Command":
                continue
            
            target_server = self.get_target_server(row, sheet, name_col, current_title_server)
            log_file = self.find_log_file(target_folder, target_server)
            
            if not log_file:
                sheet.cell(row=row, column=out_col).value = "NOT_FOUND"
                sheet.cell(row=row, column=out_col+1).value = f"server={target_server}; reason=log_file_not_found"
                continue

            input_val = sheet.cell(row=row, column=input_col).value
            val, info, cmd_ln = self.extract_from_log(log_file, str(cmd_val), options['strategy'], input_val)
            
            rel_path = log_file.relative_to(Path(log_root))
            source_text = f"path={rel_path}; {info}"
            
            sheet.cell(row=row, column=out_col).value = val if val else "NOT_FOUND"
            sheet.cell(row=row, column=out_col+1).value = source_text
            
            log_callback(f"  - [{target_server}] {cmd_val[:20]}... 처리 완료")

# ==========================================
# 2. GUI Application
# ==========================================

class ExcelAutoApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Senior Excel Log Automator v1.0")
        self.root.geometry("700x750")
        
        self.excel_path = tk.StringVar()
        self.log_dir = tk.StringVar()
        self.handlers = {"version": VersionSheetHandler()}
        
        self.create_widgets()

    def create_widgets(self):
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 1. 파일 선택
        file_frame = ttk.LabelFrame(main_frame, text="1. 파일 및 경로 설정", padding="5")
        file_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(file_frame, text="엑셀 선택", command=self.load_excel).grid(row=0, column=0, padx=2)
        ttk.Entry(file_frame, textvariable=self.excel_path, width=60).grid(row=0, column=1, padx=2)
        
        ttk.Button(file_frame, text="로그 폴더", command=self.load_log_dir).grid(row=1, column=0, padx=2, pady=5)
        ttk.Entry(file_frame, textvariable=self.log_dir, width=60).grid(row=1, column=1, padx=2)

        # 2. 시트 및 컬럼 설정
        config_frame = ttk.LabelFrame(main_frame, text="2. 시트 및 파싱 설정", padding="5")
        config_frame.pack(fill=tk.X, pady=5)

        ttk.Label(config_frame, text="대상 시트:").grid(row=0, column=0, sticky=tk.W)
        self.sheet_combo = ttk.Combobox(config_frame, state="readonly")
        self.sheet_combo.grid(row=0, column=1, sticky=tk.W, pady=2)
        
        cols_frame = ttk.Frame(config_frame)
        cols_frame.grid(row=1, column=0, columnspan=4, pady=5)
        
        self.col_vars = {}
        for i, label in enumerate(["Name", "Command", "Input Data", "Output Data"]):
            ttk.Label(cols_frame, text=f"{label}:").grid(row=0, column=i*2)
            var = tk.StringVar(value=chr(65+i))
            cb = ttk.Combobox(cols_frame, textvariable=var, values=[chr(65+j) for j in range(26)], width=3)
            cb.grid(row=0, column=i*2+1, padx=5)
            self.col_vars[label] = var

        # 3. 옵션 설정
        opt_frame = ttk.LabelFrame(main_frame, text="3. 동작 옵션", padding="5")
        opt_frame.pack(fill=tk.X, pady=5)

        ttk.Label(opt_frame, text="추출 전략:").grid(row=0, column=0, sticky=tk.W)
        self.strategy_var = tk.StringVar(value="마지막 줄")
        ttk.Radiobutton(opt_frame, text="첫 줄", variable=self.strategy_var, value="첫 줄").grid(row=0, column=1)
        ttk.Radiobutton(opt_frame, text="마지막 줄", variable=self.strategy_var, value="마지막 줄").grid(row=0, column=2)
        ttk.Radiobutton(opt_frame, text="유사도(Input 기준)", variable=self.strategy_var, value="유사도").grid(row=0, column=3)

        ttk.Label(opt_frame, text="폴더 매칭:").grid(row=1, column=0, sticky=tk.W)
        self.match_var = tk.StringVar(value="포함 매칭")
        ttk.OptionMenu(opt_frame, self.match_var, "포함 매칭", "포함 매칭", "유사도 매칭").grid(row=1, column=1, columnspan=2)

        self.overwrite_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(opt_frame, text="원본 파일 덮어쓰기", variable=self.overwrite_var).grid(row=1, column=3)

        # 4. 실행 및 로그
        self.run_btn = ttk.Button(main_frame, text="자동 입력 실행", command=self.start_process)
        self.run_btn.pack(fill=tk.X, pady=10)

        self.progress = ttk.Progressbar(main_frame, mode='determinate')
        self.progress.pack(fill=tk.X, pady=5)

        self.log_text = tk.Text(main_frame, height=15, state="disabled", font=("Consolas", 9))
        self.log_text.pack(fill=tk.BOTH, expand=True)

    def log(self, message):
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")

    def load_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.excel_path.set(path)
            wb = openpyxl.load_workbook(path, read_only=True)
            self.sheet_combo['values'] = wb.sheetnames
            if "version" in [s.lower() for s in wb.sheetnames]:
                idx = [s.lower() for s in wb.sheetnames].index("version")
                self.sheet_combo.current(idx)
            wb.close()

    def load_log_dir(self):
        path = filedialog.askdirectory()
        if path: self.log_dir.set(path)

    def start_process(self):
        if not self.excel_path.get() or not self.log_dir.get():
            messagebox.showwarning("알림", "파일과 폴더를 모두 선택해주세요.")
            return
        
        sheet_name = self.sheet_combo.get()
        if sheet_name.lower() not in self.handlers:
            messagebox.showinfo("미구현", f"'{sheet_name}' 시트 처리기는 아직 준비되지 않았습니다.")
            return

        self.run_btn.config(state="disabled")
        threading.Thread(target=self.worker, daemon=True).start()

    def worker(self):
        try:
            self.log("🚀 프로세스 시작...")
            wb = openpyxl.load_workbook(self.excel_path.get())
            sheet = wb[self.sheet_combo.get()]
            
            options = {
                'col_name': self.col_vars["Name"].get(),
                'col_cmd': self.col_vars["Command"].get(),
                'col_input': self.col_vars["Input Data"].get(),
                'col_out': self.col_vars["Output Data"].get(),
                'strategy': self.strategy_var.get(),
                'match_mode': self.match_var.get()
            }
            
            handler = self.handlers[self.sheet_combo.get().lower()]
            handler.process(sheet, self.log_dir.get(), options, self.log)
            
            # 저장
            if self.overwrite_var.get():
                save_path = self.excel_path.get()
            else:
                p = Path(self.excel_path.get())
                save_path = str(p.parent / f"{p.stem}_updated{p.suffix}")
            
            wb.save(save_path)
            self.log(f"✅ 작업 완료! 저장 위치: {save_path}")
            messagebox.showinfo("성공", f"파일이 성공적으로 저장되었습니다.\n{save_path}")
            
        except Exception as e:
            self.log(f"🔥 치명적 오류 발생: {str(e)}")
            messagebox.showerror("오류", str(e))
        finally:
            self.run_btn.config(state="normal")
            self.progress['value'] = 0

if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelAutoApp(root)
    root.mainloop()